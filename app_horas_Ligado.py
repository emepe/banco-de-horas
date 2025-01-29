import pandas as pd
import os
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font

arquivo = r"C:\Users\Maria.Paula\Desktop\Projeto1_AppHoras\Horários_de_entrada_x_saída.xlsx"

horario_atual = datetime.now()
data_formatada = horario_atual.strftime("%d/%m/%Y")
hora_formatada = horario_atual.strftime("%H:%M:%S")


novo_registro = {
    "Data": [data_formatada],
    "Hora": [hora_formatada],
    "Evento": ["Ligado"]
}

df_novo = pd.DataFrame(novo_registro)

if os.path.exists(arquivo):

        df_existente = pd.read_excel(arquivo)
        df_atualizado = pd.concat([df_existente, df_novo], ignore_index=True)

else:

        df_atualizado = df_novo

df_atualizado.to_excel(arquivo, index=False)

workbook = load_workbook(arquivo)
sheet = workbook.active

for column in sheet.columns:
        max_lenght = 0 
        column_letter = get_column_letter(column[0].column)
        for cell in column:
                if cell.value:
                        max_lenght = max(max_lenght, len(str(cell.value)))
        adjusted_width = max_lenght + 2 
        sheet.column_dimensions[column_letter].width = adjusted_width


for cell in sheet [1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

workbook.save(arquivo)

print(f"Evento 'Ligado' registrado com sucesso.")

